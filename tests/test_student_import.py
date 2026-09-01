"""Import tests use an isolated in-memory database and never send real emails."""
import unittest
import io
from zipfile import ZipFile
from unittest.mock import AsyncMock, patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, select, func
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from app.core.database import get_db
from app.core.errors import APIError, ConflictError, ValidationError, api_error_handler
from app.modules.auth.dependencies import get_current_user
from app.modules.auth.models import AccessLevel, User, UserAccessLevel
from app.modules.auth.schemas import CurrentUser
from app.modules.lms.models import StudentProfile
from app.modules.lms.router import router
from app.modules.lms import student_import_service as service
from app.modules.lms.student_excel_import import parse_excel_students
from app.modules.lms.schemas.student_import import StudentImportRequest

HEADER = "full_name,email,student_number,phone,notes\n"
CSV = HEADER + ' Alice Example ,ALICE@example.com,stu-01,+94770000000,"A note, with comma"\nBob Example,bob@example.com,stu-02,,\n'


class ImportParsingTests(unittest.TestCase):
    def test_csv_bom_unicode_quotes_newlines_and_normalization(self):
        rows = service.parse_students('\ufeff' + HEADER + '"Doe, Ana",ANA@example.com,stu-001,001234,"First line\nSecond line"\n\nඅමාලි,amali@example.com,stu-002,,\n')
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].full_name, "Doe, Ana")
        self.assertEqual(rows[0].email, "ana@example.com")
        self.assertEqual(rows[0].student_number, "STU-001")
        self.assertEqual(rows[0].phone, "001234")
        self.assertEqual(rows[0].notes, "First line\nSecond line")
        self.assertEqual(rows[1].full_name, "අමාලි")
        self.assertFalse(any(row.errors for row in rows))

    def test_required_fields_invalid_email_length_and_duplicate_rows(self):
        rows = service.parse_students(HEADER + ' ,invalid, ,,' + 'x' * 5001 + '\nA,a@example.com,abc,,\nB,A@example.com,ABC,,\n')
        self.assertGreaterEqual(len(rows[0].errors), 4)
        for row in rows[1:]:
            self.assertEqual(len(row.errors), 2)
            self.assertTrue(all("repeated" in error for error in row.errors))

    def test_column_mismatch_is_reported_on_row(self):
        row = service.parse_students(HEADER + "A,a@example.com,one,123,unquoted,comma\n")[0]
        self.assertIn("Column count", row.errors[0])

    def test_rejects_unsupported_headers_broken_csv_empty_and_oversized_batches(self):
        for text in ["", HEADER, "name,email\nA,a@example.com", "full_name,email,email,student_number\nA,a,b,c",
                     HEADER.strip() + ',access\nA,a@example.com,S1,,,ADMIN', HEADER + '"unfinished', HEADER + '\x00',
                     HEADER + "x" * 500_001, HEADER + "A,a@example.com,S1,,\n" * 101]:
            with self.subTest(text=text[:60]):
                with self.assertRaises(ValidationError):
                    service.parse_students(text)


class ExcelParsingTests(unittest.TestCase):
    @staticmethod
    def workbook_bytes(configure=None):
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Students"
        sheet.append(["full_name", "email", "student_number", "phone", "notes"])
        sheet.append(["Ana Perera", "ANA@example.com", "STU-001", 1234, "Excel row"])
        sheet["D2"].number_format = "0000000000"
        if configure:
            configure(workbook, sheet)
        data = io.BytesIO()
        workbook.save(data)
        workbook.close()
        return data.getvalue()

    def test_reads_first_worksheet_values_and_preserves_displayed_zeroes(self):
        rows, sheet = parse_excel_students(self.workbook_bytes(lambda workbook, current: workbook.create_sheet("Ignored")))
        self.assertEqual(sheet, "Students")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].row, 2)
        self.assertEqual(rows[0].email, "ana@example.com")
        self.assertEqual(rows[0].phone, "0000001234")
        self.assertFalse(rows[0].errors)

    def test_formula_merge_hidden_sheet_and_unsupported_cell_values_are_rejected(self):
        mutations = [
            lambda workbook, sheet: setattr(sheet["E2"], "value", "=1+1"),
            lambda workbook, sheet: sheet.merge_cells("A1:B1"),
            lambda workbook, sheet: (workbook.create_sheet("Visible"), setattr(sheet, "sheet_state", "hidden")),
            lambda workbook, sheet: setattr(sheet["D2"], "value", 123.5),
            lambda workbook, sheet: setattr(sheet["D2"], "value", True),
        ]
        for mutation in mutations:
            with self.subTest(mutation=mutation):
                with self.assertRaises(ValidationError):
                    parse_excel_students(self.workbook_bytes(mutation))

    def test_extra_columns_rows_macros_invalid_files_and_archive_bombs_are_rejected(self):
        mutations = [
            lambda workbook, sheet: setattr(sheet["F2"], "value", "unexpected"),
            lambda workbook, sheet: setattr(sheet["A1002"], "value", "too far"),
        ]
        for mutation in mutations:
            with self.assertRaises(ValidationError):
                parse_excel_students(self.workbook_bytes(mutation))
        for data in [b"not xlsx", b"", b"x" * 2_000_001]:
            with self.assertRaises(ValidationError):
                parse_excel_students(data)
        bomb = io.BytesIO()
        with ZipFile(bomb, "w") as archive:
            archive.writestr("large.xml", b"x" * 8_000_001)
        with self.assertRaises(ValidationError):
            parse_excel_students(bomb.getvalue())

    def test_xml_entities_are_rejected_before_workbook_parsing(self):
        data = io.BytesIO()
        with ZipFile(data, "w") as archive:
            archive.writestr("[Content_Types].xml", b'<!DOCTYPE x [<!ENTITY y "boom">]><x>&y;</x>')
        with self.assertRaises(ValidationError):
            parse_excel_students(data.getvalue())
        data = io.BytesIO()
        with ZipFile(data, "w") as archive:
            archive.writestr("[Content_Types].xml", b"<x>macroEnabled</x>")
        with self.assertRaises(ValidationError):
            parse_excel_students(data.getvalue())


class LocalSession:
    def __init__(self, session): self.session = session
    def add_all(self, values): self.session.add_all(values)
    async def execute(self, statement): return self.session.execute(statement)
    async def flush(self): self.session.flush()
    async def commit(self): self.session.commit()
    async def rollback(self): self.session.rollback()


class ImportDatabaseTests(unittest.IsolatedAsyncioTestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        for model in [User, AccessLevel, UserAccessLevel, StudentProfile]:
            model.__table__.create(self.engine)
        self.session = Session(self.engine, expire_on_commit=False)
        self.db = LocalSession(self.session)
        self.session.add(User(user_id=1, email="admin@example.com", full_name="Admin"))
        self.session.add_all([AccessLevel(access_key=key, display_name=key) for key in ["LMS", "STUDENT", "ADMIN"]])
        self.session.commit()
        self.addCleanup(self.engine.dispose)
        self.addCleanup(self.session.close)

    def count(self, model): return self.session.scalar(select(func.count()).select_from(model))

    async def test_preview_then_create_once_with_only_student_access(self):
        preview = await service.import_students(self.db, StudentImportRequest(csv_text=CSV), 1)
        self.assertTrue(preview.can_import)
        self.assertEqual(preview.imported, 0)
        self.assertEqual(self.count(User), 1)
        result = await service.import_students(self.db, StudentImportRequest(csv_text=CSV, preview=False), 1)
        self.assertEqual(result.imported, 2)
        self.assertFalse(result.can_import)
        self.assertEqual(self.count(User), 3)
        self.assertEqual(self.count(StudentProfile), 2)
        for row in result.rows:
            user = self.session.get(User, row.user_id)
            self.assertEqual({grant.access_level.access_key for grant in user.access_levels}, {"LMS", "STUDENT"})
            self.assertEqual(user.created_by, 1)
            self.assertEqual(self.session.get(StudentProfile, row.user_id).student_number, row.student_number)
        retry = await service.import_students(self.db, StudentImportRequest(csv_text=CSV, preview=False), 1)
        self.assertFalse(retry.can_import)
        self.assertEqual(retry.imported, 0)
        self.assertEqual(self.count(User), 3)

    async def test_any_invalid_row_blocks_entire_batch(self):
        payload = StudentImportRequest(csv_text=CSV + 'Invalid,not-an-email,STU-03,,\n', preview=False)
        result = await service.import_students(self.db, payload, 1)
        self.assertFalse(result.can_import)
        self.assertEqual(self.count(User), 1)
        self.assertEqual(self.count(StudentProfile), 0)

    async def test_maximum_batch_commits_once(self):
        text = HEADER + "".join(f"Student {index},s{index}@example.com,STU-{index},,\n" for index in range(100))
        with patch.object(self.db, "commit", wraps=self.db.commit) as commit:
            result = await service.import_students(self.db, StudentImportRequest(csv_text=text, preview=False), 1)
        self.assertEqual(result.imported, 100)
        self.assertEqual(len({row.user_id for row in result.rows}), 100)
        self.assertEqual(self.count(StudentProfile), 100)
        commit.assert_awaited_once()

    async def test_conflicts_across_all_accounts_case_insensitive_and_rechecked_at_confirm(self):
        await service.import_students(self.db, StudentImportRequest(csv_text=CSV), 1)
        self.session.add(User(user_id=2, email="ALICE@EXAMPLE.COM", full_name="Existing CMS account"))
        self.session.add(StudentProfile(user_id=2, student_number="stu-02"))
        self.session.commit()
        result = await service.import_students(self.db, StudentImportRequest(csv_text=CSV, preview=False), 1)
        self.assertIn("Email already", result.rows[0].errors[0])
        self.assertIn("Student number already", result.rows[1].errors[0])
        self.assertEqual(self.count(User), 2)
        self.assertEqual(self.session.get(User, 2).full_name, "Existing CMS account")

    async def test_inactive_access_cannot_create_unusable_accounts(self):
        self.session.scalar(select(AccessLevel).where(AccessLevel.access_key == "STUDENT")).is_active = False
        self.session.commit()
        with self.assertRaises(ValidationError):
            await service.import_students(self.db, StudentImportRequest(csv_text=CSV, preview=False), 1)
        self.assertEqual(self.count(User), 1)

    async def test_mid_transaction_constraint_failure_rolls_back_all_new_users(self):
        async def fail_commit():
            self.session.flush()  # Users, profiles and grants are now staged, but not committed.
            raise IntegrityError("test", {}, Exception("simulated concurrent duplicate"))
        with patch.object(self.db, "commit", side_effect=fail_commit):
            with self.assertRaises(ConflictError):
                await service.import_students(self.db, StudentImportRequest(csv_text=CSV, preview=False), 1)
        self.assertEqual(self.count(User), 1)
        self.assertEqual(self.count(StudentProfile), 0)
        self.assertEqual(self.count(UserAccessLevel), 0)


class ImportAccessTests(unittest.TestCase):
    def setUp(self):
        self.app = FastAPI()
        self.app.add_exception_handler(APIError, api_error_handler)
        self.app.include_router(router)
        self.app.dependency_overrides[get_db] = lambda: None
        self.client = TestClient(self.app)
        self.addCleanup(self.client.close)

    def user(self, access):
        self.app.dependency_overrides[get_current_user] = lambda: CurrentUser(user_id=9, email="admin@example.com", access=access)

    def test_import_and_preview_are_admin_only(self):
        self.assertEqual(self.client.post("/api/v1/lms/students/import", json={"csv_text": CSV}).status_code, 401)
        for access in [["LMS", "STUDENT"], ["LMS", "LECTURER"], ["CMS", "ADMIN"], ["LMS"]]:
            self.user(access)
            for preview in [True, False]:
                with patch.object(service, "import_students", new_callable=AsyncMock) as importer:
                    response = self.client.post("/api/v1/lms/students/import", json={"csv_text": CSV, "preview": preview})
                    self.assertEqual(response.status_code, 403)
                    importer.assert_not_awaited()

    def test_admin_roles_can_preview_with_no_store_and_server_identity(self):
        for role in ["ADMIN", "SUPER_ADMIN"]:
            self.user(["LMS", role])
            with patch.object(service, "import_students", new_callable=AsyncMock, return_value={"rows": [], "can_import": False, "imported": 0}) as importer:
                response = self.client.post("/api/v1/lms/students/import", json={"csv_text": CSV})
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.headers["cache-control"], "no-store")
                self.assertTrue(importer.await_args.args[1].preview)
                self.assertEqual(importer.await_args.args[2], 9)

    def test_excel_endpoint_has_same_admin_guard_and_validates_media_type(self):
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        self.user(["LMS", "STUDENT"])
        with patch("app.modules.lms.router.parse_excel_students") as parser:
            self.assertEqual(self.client.post("/api/v1/lms/students/import/excel", content=b"xlsx", headers={"Content-Type": content_type}).status_code, 403)
            parser.assert_not_called()
        self.user(["LMS", "ADMIN"])
        self.assertEqual(self.client.post("/api/v1/lms/students/import/excel", content=b"csv", headers={"Content-Type": "text/csv"}).status_code, 415)

    def test_admin_can_preview_excel_and_sheet_name_is_returned(self):
        from app.modules.lms.schemas.student_import import StudentImportRow, StudentImportResponse
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        self.user(["LMS", "ADMIN"])
        rows = [StudentImportRow(row=2, full_name="A", email="a@example.com", student_number="S1")]
        with (
            patch("app.modules.lms.router.parse_excel_students", return_value=(rows, "Students")) as parser,
            patch.object(service, "import_student_rows", new_callable=AsyncMock,
                         return_value=StudentImportResponse(rows=rows, can_import=True)) as importer,
        ):
            response = self.client.post("/api/v1/lms/students/import/excel?preview=true", content=b"xlsx", headers={"Content-Type": content_type})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["sheet_name"], "Students")
        self.assertEqual(response.headers["cache-control"], "no-store")
        parser.assert_called_once_with(b"xlsx")
        self.assertTrue(importer.await_args.args[2])
        self.assertEqual(importer.await_args.args[3], 9)

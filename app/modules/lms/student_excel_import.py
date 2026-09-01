"""Read bounded, values-only Excel uploads; reuse the CSV validation/creation flow."""
import csv
import io
import math
import re
from datetime import date, datetime, time, timedelta
from zipfile import ZipFile, BadZipFile

from defusedxml.ElementTree import fromstring
from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from xml.etree.ElementTree import ParseError

from app.core.errors import ValidationError
from app.modules.lms.student_import_service import parse_students

MAX_EXCEL_BYTES = 2_000_000
MAX_EXPANDED_BYTES = 8_000_000


def _check_archive(data: bytes):
    if not data or len(data) > MAX_EXCEL_BYTES:
        raise ValidationError("Excel files must be .xlsx and at most 2 MB.")
    with ZipFile(io.BytesIO(data)) as archive:
        entries = archive.infolist()
        if len(entries) > 200 or sum(entry.file_size for entry in entries) > MAX_EXPANDED_BYTES:
            raise ValidationError("This workbook is too complex. Copy the student table to a new .xlsx workbook.")
        if len({entry.filename for entry in entries}) != len(entries):
            raise ValidationError("This workbook contains duplicate archive entries. Save a fresh .xlsx copy.")
        for entry in entries:
            if entry.flag_bits & 1 or entry.filename.lower().endswith(".bin"):
                raise ValidationError("Encrypted or macro-enabled workbooks are not supported. Save as a plain .xlsx file.")
            if entry.filename.endswith((".xml", ".rels")):
                # Validate before passing any XML to the workbook parser. No DTDs,
                # entities or external entity expansion, even if library flags differ.
                xml = archive.read(entry)
                if b"macroEnabled" in xml:
                    raise ValidationError("Macro-enabled workbooks are not supported. Save as a plain .xlsx file.")
                fromstring(xml, forbid_dtd=True, forbid_entities=True, forbid_external=True)


def _cell_text(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if cell.data_type in {"f", "e"}:
        raise ValidationError(f"Cell {cell.coordinate}: formulas and Excel errors are not supported. Paste values only.")
    if isinstance(value, (bool, datetime, date, time, timedelta)):
        raise ValidationError(f"Cell {cell.coordinate}: use text, not dates, times or true/false values.")
    if isinstance(value, (int, float)):
        if not math.isfinite(value) or value != int(value) or abs(value) >= 10**15:
            raise ValidationError(f"Cell {cell.coordinate}: enter this value as Text to avoid rounding or lost digits.")
        # Preserve common zero-padded identifier/phone formats as displayed in Excel.
        if re.fullmatch(r"0{1,50}", cell.number_format or "") and value >= 0:
            return str(int(value)).zfill(len(cell.number_format))
        return str(int(value))
    return str(value)


def parse_excel_students(data: bytes):
    workbook = None
    try:
        _check_archive(data)
        # Normal mode determines actual populated bounds instead of trusting stale
        # worksheet dimensions. Archive expansion is capped above; parsing runs off
        # the API event loop and never evaluates formulas or follows external links.
        workbook = load_workbook(io.BytesIO(data), read_only=False, data_only=False, keep_links=False, keep_vba=False)
        if not workbook.worksheets:
            raise ValidationError("The workbook has no worksheets.")
        sheet = workbook.worksheets[0]
        if sheet.sheet_state != "visible" or sheet.merged_cells.ranges:
            raise ValidationError("The first worksheet must be visible and contain a plain table without merged cells.")
        if sheet.max_row > 1001 or sheet.max_column > 5:
            raise ValidationError("Use only columns A–E and at most 100 student rows. Remove extra formatted rows/columns or copy the table to a new workbook.")
        text = io.StringIO(newline="")
        writer = csv.writer(text)
        source_rows = []
        for index, cells in enumerate(sheet.iter_rows(), start=1):
            values = [_cell_text(cell) for cell in cells]
            if index > 1 and not any(value.strip() for value in values):
                continue
            if index == 1:
                while values and not values[-1].strip():
                    values.pop()
                width = len(values)
            else:
                if any(value.strip() for value in values[width:]):
                    raise ValidationError(f"Row {index}: values appear below an empty column header.")
                values = values[:width]
                source_rows.append(index)
            writer.writerow(values)
        rows = parse_students(text.getvalue())
        for row, source_row in zip(rows, source_rows):
            row.row = source_row
            row.errors = [error.replace("this CSV", "this worksheet") for error in row.errors]
        return rows, sheet.title
    except ValidationError:
        raise
    except (BadZipFile, InvalidFileException, ParseError, DefusedXmlException, ValueError, KeyError, TypeError, OSError, RuntimeError, OverflowError):
        raise ValidationError("Could not read this Excel file. Use an unprotected .xlsx workbook; .xls and .xlsm are not supported.") from None
    finally:
        if workbook is not None:
            workbook.close()
